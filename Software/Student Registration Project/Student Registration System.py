from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl
import pathlib

from openpyxl.workbook import Workbook

background = '#1A5276'
framebg = 'cyan'
framefg = 'black'

root = Tk()
root.title('Student Registration System')
root.geometry('800x600+210+70')
root.config(bg=background)

file = pathlib.Path('Student Information.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Semester"
    sheet['D1'] = "Gender"
    sheet['E1'] = "DOB"
    sheet['F1'] = "Date of Registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Contact"
    sheet['J1'] = "Dept."
    sheet['K1'] = "Father's Name"
    sheet['L1'] = "Father's Occupation"
    sheet['M1'] = "Father's Contact"
    sheet['N1'] = "Mother's Name"
    sheet['O1'] = "Mother's Occupation"
    sheet['P1'] = "Mother's Contact"
    file.save('Student Information.xlsx')


# Exit
def Exit():
    root.destroy()


# Image Viewer
def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select Image File", filetypes=(('JPG File', "*.jpg"),
                                                                                ("PNG File", "*.png"),
                                                                                ("All Files", "*.txt")))
    img = (Image.open(filename))
    resized_img = img.resize((150, 150))
    photo2 = ImageTk.PhotoImage(resized_img)
    lbl.config(image=photo2)
    lbl.image = photo2


# Auto Registration No. updater
def registration_no():
    file = openpyxl.load_workbook('Student Information.xlsx')
    sheet = file.active
    row = sheet.max_row
    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set('1')


# Clear
def Clear():
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Contact.set('')
    Skills.set('')
    Father_Name.set('')
    Father_Contact.set('')
    Father_Occu.set('')
    Mother_Name.set('')
    Mother_Occu.set('')
    Mother_Contact.set('')
    Save_Button.config(state='normal')
    Sem.set('Select Semester')
    DEPT.set('Choose Department')

    img1 = PhotoImage(file='Images/2.png')
    lbl.config(image=img1)
    lbl.image = img1
    img = ""


# save
def Save():
    R1 = Registration.get()
    N1 = Name.get()
    S1 = Sem.get()
    try:
        G1 = gender
    except:
        messagebox.showerror("Error!", "Select Gender.")
    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    Sk1 = Skills.get()
    Con = Contact.get()
    Dept = DEPT.get()
    fname = Father_Name.get()
    fatheroccu = Father_Occu.get()
    fathercontact = Father_Contact.get()
    mname = Mother_Name.get()
    motheroccu = Mother_Occu.get()
    mothercontact = Mother_Contact.get()

    if R1 == '' or N1 == '' or S1 == '' or G1 == '' or Dept == '' or D1 == '' or D2 == '' or Re1 == '' or Sk1 == '' or Con == '' or fname == '' or mname == '' or fatheroccu == '' or fathercontact == '' or motheroccu == '' or mothercontact == '':
        messagebox.showerror('Error!', 'Few data missing.')
    else:
        file = openpyxl.load_workbook('Student Information.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=S1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row, value=D2)
        sheet.cell(column=6, row=sheet.max_row, value=D1)
        sheet.cell(column=7, row=sheet.max_row, value=Re1)
        sheet.cell(column=8, row=sheet.max_row, value=Sk1)
        sheet.cell(column=9, row=sheet.max_row, value=Con)
        sheet.cell(column=10, row=sheet.max_row, value=Dept)
        sheet.cell(column=11, row=sheet.max_row, value=fname)
        sheet.cell(column=12, row=sheet.max_row, value=fatheroccu)
        sheet.cell(column=13, row=sheet.max_row, value=fathercontact)
        sheet.cell(column=14, row=sheet.max_row, value=mname)
        sheet.cell(column=15, row=sheet.max_row, value=motheroccu)
        sheet.cell(column=16, row=sheet.max_row, value=mothercontact)
        file.save(r'Student Information.xlsx')

        try:
            img.save('Student Images/' + str(R1) + '.jpg')
        except:
            messagebox.showinfo("Info", "Profile picture not available")

        messagebox.showinfo("Info", "Successfully data Entered.")

        Clear()

        registration_no()


# Search
def search():
    text = Search.get()
    Clear()
    Save_Button.config(state='disabled')

    file = openpyxl.load_workbook('Student Information.xlsx')
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]

    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid!", "Invalid registration Number.")

    x1 = sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value
    x13 = sheet.cell(row=int(reg_number), column=13).value
    x14 = sheet.cell(row=int(reg_number), column=14).value
    x15 = sheet.cell(row=int(reg_number), column=15).value
    x16 = sheet.cell(row=int(reg_number), column=16).value

    Registration.set(x1)
    Name.set(x2)
    Sem.set(x3)
    if x4 == 'Female':
        R2.select()
    elif x4 == 'Male':
        R1.select()
    else:
        R3.select()
    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skills.set(x8)
    Contact.set(x9)
    DEPT.set(x10)
    Father_Name.set(x11)
    Father_Occu.set(x12)
    Father_Contact.set(x13)
    Mother_Name.set(x14)
    Mother_Occu.set(x15)
    Mother_Contact.set(x16)

    img = (Image.open("Student Images/" + str(x1) + ".jpg"))
    resized_image = img.resize((150, 150))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2


# update
def Update():
    R1 = Registration.get()
    N1 = Name.get()
    S1 = Sem.get()
    select()
    G1 = gender
    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    Sk1 = Skills.get()
    Con = Contact.get()
    Dept = DEPT.get()
    fname = Father_Name.get()
    fatheroccu = Father_Occu.get()
    fathercontact = Father_Contact.get()
    mname = Mother_Name.get()
    motheroccu = Mother_Occu.get()
    mothercontact = Mother_Contact.get()

    file = openpyxl.load_workbook('Student Information.xlsx')
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == R1:
            name = row[0]
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]

    sheet.cell(column=2, row=int(reg_number), value=N1)
    sheet.cell(column=3, row=int(reg_number), value=S1)
    sheet.cell(column=4, row=int(reg_number), value=G1)
    sheet.cell(column=5, row=int(reg_number), value=D2)
    sheet.cell(column=6, row=int(reg_number), value=D1)
    sheet.cell(column=7, row=int(reg_number), value=Re1)
    sheet.cell(column=8, row=int(reg_number), value=Sk1)
    sheet.cell(column=9, row=int(reg_number), value=Con)
    sheet.cell(column=10, row=int(reg_number), value=Dept)
    sheet.cell(column=11, row=int(reg_number), value=fname)
    sheet.cell(column=12, row=int(reg_number), value=fatheroccu)
    sheet.cell(column=13, row=int(reg_number), value=fathercontact)
    sheet.cell(column=14, row=int(reg_number), value=mname)
    sheet.cell(column=15, row=int(reg_number), value=motheroccu)
    sheet.cell(column=16, row=int(reg_number), value=mothercontact)

    file.save(r'Student Information.xlsx')

    try:
        img.save("Student Images/" + str(R1) + ".jpg")
    except:
        pass

    messagebox.showinfo("Update", "Updated Successfully!")

    Clear()


# gender
def select():
    global gender
    value = radio.get()
    if value == 1:
        gender = 'Male'
    elif value == 2:
        gender = 'Female'
    else:
        gender = 'Other'


# top frames
Label(root, text='Project 1, created by Sakhawat Hossain Mahin', width=10, height=1,
      bg='cyan', fg='black', anchor='w').pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=2, bg='#CF6B08', fg='white',
      font='arial 20 bold').pack(side=TOP, fill=X)

# search box to update
Search = StringVar()
Entry(root, textvariable=Search, width=20, bd=2, font='arial 10').place(x=580, y=44)
imageicon3 = PhotoImage(file='Images/search.gif')
Srch = Button(root, text='Search', compound=LEFT, width=5, height=1, bd=2, bg='black', fg='white', font='arial 10 bold',
              command=search)
Srch.place(x=730, y=41)

imageicon4 = PhotoImage(file='Images/profile.png')
Update_button = Button(root, text='UPDATE', bg='Green', fg='white', bd=2, font='arial 14 bold', command=Update)
Update_button.place(x=70, y=37)

# registration and date
Label(root, text='Registration No.', font='arial 10 bold', fg='white', bg=background).place(x=30, y=130)
Label(root, text='Date: ', font='arial 10 bold', fg='white', bg=background).place(x=210, y=130)
Label(root, text='Session: ', font='arial 10 bold', fg='white', bg=background).place(x=424, y=130)

Registration = IntVar()
Date = StringVar()
Session = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=5, font='arial 10')
reg_entry.place(x=140, y=130)

registration_no()

today = date.today()
d1 = today.strftime('%d-%m-%y')
date_entry = Entry(root, textvariable=Date, width=20, font='arial 10 italic')
date_entry.place(x=250, y=130)
Date.set(d1)

year = date.today()
yr1 = year.strftime('%y')
year_entry = Entry(root, textvariable=Session, width=20, font='arial 10 italic')
year_entry.place(x=485, y=130)
Session.set(yr1)

# Students Details
obj = LabelFrame(root, text='Personal Information', font='arial 10 italic', bd=1, width=600, height=185,
                 bg=background, fg='white', relief=GROOVE)
obj.place(x=30, y=180)

Label(obj, text='Full Name:', font='arial 10', bg=background, fg='white').place(x=25, y=20)
Label(obj, text='Date of Birth:', font='arial 10', bg=background, fg='white').place(x=25, y=50)
Label(obj, text='Gender:', font='arial 10', bg=background, fg='white').place(x=25, y=80)
Label(obj, text='Dept.:', font='arial 10', bg=background, fg='white').place(x=25, y=110)

Label(obj, text='Semester:', font='arial 10', bg=background, fg='white').place(x=320, y=20)
Label(obj, text='Religion:', font='arial 10', bg=background, fg='white').place(x=320, y=50)
Label(obj, text='Skills:', font='arial 10', bg=background, fg='white').place(x=320, y=80)
Label(obj, text='Contact:', font='arial 10', bg=background, fg='white').place(x=320, y=110)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=30, font='arial 10')
name_entry.place(x=95, y=20)

DOB = StringVar()
dob_entry = Entry(obj, textvariable=DOB, width=20, font='arial 10')
dob_entry.place(x=110, y=50)

radio = IntVar()
R1 = Radiobutton(obj, text='Male', variable=radio, value=1, bg=background, fg='red', command=select())
R1.place(x=80, y=80)
R2 = Radiobutton(obj, text='Female', variable=radio, value=2, bg=background, fg='red', command=select())
R2.place(x=140, y=80)
R3 = Radiobutton(obj, text='Other', variable=radio, value=3, bg=background, fg='red', command=select())
R3.place(x=210, y=80)

DEPT = Combobox(obj, values=['Computer Science & Engineering (CSE)',
                             'Computer Science & Information Technology (CSIT)',
                             'Fashion Design & Technology (FDT)',
                             'Interior Architecture (IA)',
                             'Bachelor of Business Administration (BBA)',
                             'Fine Arts (FA)'], font='roboto 10', width=30, state='r')
DEPT.place(x=70, y=110)
DEPT.set('Choose Department')

Sem = Combobox(obj, values=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'], font='roboto 10', width=17,
               state='r', )
Sem.place(x=390, y=20)
Sem.set('Select Semester')

Religion = StringVar()
religion_entry = Entry(obj, textvariable=Religion, width=20, font='arial 10')
religion_entry.place(x=390, y=50)

Skills = StringVar()
skill_entry = Entry(obj, textvariable=Skills, width=20, font='arial 10')
skill_entry.place(x=390, y=80)

Contact = StringVar()
contact_entry = Entry(obj, textvariable=Contact, width=20, font='arial 10')
contact_entry.place(x=390, y=110)

# Parent's Details
obj2 = LabelFrame(root, text="Parent's Information", font='arial 10 italic', bd=1, width=600, height=185, bg=background,
                  fg='white',
                  relief=GROOVE)
obj2.place(x=30, y=390)

Label(obj2, text="Father's Name:", font='arial 10', bg=background, fg='white').place(x=25, y=20)
Label(obj2, text="Occupation:", font='arial 10', bg=background, fg='white').place(x=25, y=70)
Label(obj2, text="Contact:", font='arial 10', bg=background, fg='white').place(x=25, y=120)

Label(obj2, text="Mother's Name:", font='arial 10', bg=background, fg='white').place(x=310, y=20)
Label(obj2, text="Occupation:", font='arial 10', bg=background, fg='white').place(x=310, y=70)
Label(obj2, text="Contact:", font='arial 10', bg=background, fg='white').place(x=310, y=120)

Father_Name = StringVar()
fathername_entry = Entry(obj2, textvariable=Father_Name, width=25, font='arial 10')
fathername_entry.place(x=120, y=20)

Father_Occu = StringVar()
fatheroccu_entry = Entry(obj2, textvariable=Father_Occu, width=25, font='arial 10')
fatheroccu_entry.place(x=120, y=70)

Father_Contact = StringVar()
fathercontact_entry = Entry(obj2, textvariable=Father_Contact, width=25, font='arial 10')
fathercontact_entry.place(x=120, y=120)

Mother_Name = StringVar()
mothername_entry = Entry(obj2, textvariable=Mother_Name, width=25, font='arial 10')
mothername_entry.place(x=410, y=20)

Mother_Occu = StringVar()
motheroccu_entry = Entry(obj2, textvariable=Mother_Occu, width=25, font='arial 10')
motheroccu_entry.place(x=410, y=70)

Mother_Contact = StringVar()
mothercontact_entry = Entry(obj2, textvariable=Mother_Contact, width=25, font='arial 10')
mothercontact_entry.place(x=410, y=120)

# image
f = Frame(root, bd=3, bg='black', width=150, height=150, relief=GROOVE)
f.place(x=640, y=188)

img = PhotoImage(file='Images/2.png')
lbl = Label(f, bg='black', image=img)
lbl.place(x=0, y=0)

# buttons
Upload_Button = Button(root, text='Upload', width=17, height=1, bd=2, font='airal 10 bold', bg='#F29336',
                       command=showimage)
Upload_Button.place(x=643, y=345)

Save_Button = Button(root, text='Save', width=17, height=1, bd=2, font='airal 10 bold', bg='#F29336', command=Save)
Save_Button.place(x=643, y=385)

Reset_Button = Button(root, text='Reset', width=17, height=1, bd=2, font='airal 10 bold', bg='#F29336', command=Clear)
Reset_Button.place(x=643, y=425)

Exit_Button = Button(root, text='Exit', width=17, height=1, bd=2, font='airal 10 bold', bg='red', command=Exit)
Exit_Button.place(x=643, y=465)

root.mainloop()
